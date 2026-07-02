import os
import random
import pandas as pd
from openpyxl import Workbook

output_dir = os.path.join(os.path.dirname(__file__), 'dummy_sheets')
os.makedirs(output_dir, exist_ok=True)

def generate_sheet1():
    wb = Workbook()
    ws = wb.active
    ws.append(["Matric Number", "Name", "Course Code", "Score", "Grade"])
    students = [
        ("CSC/21/001", "Alice Johnson"),
        ("CSC/21/002", "Bob Smith"),
        ("CSC/21/003", "Charlie Davis"),
        ("CSC/21/004", "Diana Prince"),
        ("CSC/21/005", "Evan Wright"),
    ]
    courses = ["CSC401", "CSC402", "CSC403", "MTH213"]
    
    for mat, name in students:
        for course in courses:
            score = random.randint(40, 100)
            if score >= 70: grade = 'A'
            elif score >= 60: grade = 'B'
            elif score >= 50: grade = 'C'
            elif score >= 45: grade = 'D'
            else: grade = 'F'
            ws.append([mat, name, course, score, grade])
    
    path = os.path.join(output_dir, 'simple_long_format.xlsx')
    wb.save(path)
    return path

def generate_sheet2():
    wb = Workbook()
    ws = wb.active
    # First row
    ws.append(["S/N", "Matric Number", "Sex", "CSC401", "CSC402", "CSC403", "MTH213"])
    # Second row
    ws.append(["", "", "", 3, 2, 3, 3])
    # Third row
    ws.append(["", "", "", "C", "E", "C", "C"])
    
    # 15 data rows
    for i in range(1, 16):
        matric = f"FOS/22/23/{100000 + i}"
        sex = random.choice(["M", "F"])
        row = [i, matric, sex]
        for _ in range(4):
            if random.random() < 0.2: # 20% chance of blank
                row.append("")
            else:
                score = random.randint(40, 100)
                if score >= 70: grade = 'A'
                elif score >= 60: grade = 'B'
                elif score >= 50: grade = 'C'
                elif score >= 45: grade = 'D'
                else: grade = 'F'
                row.append(f"{score}{grade}")
        ws.append(row)
        
    path = os.path.join(output_dir, 'wide_broadsheet_format.xlsx')
    wb.save(path)
    return path

def generate_sheet3():
    wb = Workbook()
    ws = wb.active
    ws.append(["S/N", "Reg No", "Full Name", "Course", "CA (30)", "Exam (70)", "Grade"])
    
    courses = ["PHY101", "CHM101", "BIO101", "MTH101"]
    for i in range(1, 16):
        reg_no = f"SCI/23/{200 + i}"
        name = f"Student {i}"
        course = random.choice(courses)
        ca = random.randint(10, 30)
        exam = random.randint(20, 70)
        total = ca + exam
        if total >= 70: grade = 'A'
        elif total >= 60: grade = 'B'
        elif total >= 50: grade = 'C'
        elif total >= 45: grade = 'D'
        else: grade = 'F'
        ws.append([i, reg_no, name, course, ca, exam, grade])
        
    path = os.path.join(output_dir, 'inconsistent_columns.xlsx')
    wb.save(path)
    return path

if __name__ == "__main__":
    p1 = generate_sheet1()
    p2 = generate_sheet2()
    p3 = generate_sheet3()
    
    print("Files generated successfully.")
    
    df1 = pd.read_excel(p1, header=None)
    df2 = pd.read_excel(p2, header=None)
    df3 = pd.read_excel(p3, header=None)
    
    print(f"simple_long_format.xlsx shape (rows, cols): {df1.shape}")
    print(f"wide_broadsheet_format.xlsx shape (rows, cols): {df2.shape}")
    print(f"inconsistent_columns.xlsx shape (rows, cols): {df3.shape}")
